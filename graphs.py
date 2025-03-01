import openpyxl
import pandas as pd
import streamlit as st
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.plotarea import DataTable
from streamlit_option_menu import option_menu
import win32com.client
import PIL
from PIL import ImageGrab, Image
import os
import sys
import pythoncom


def convert_to_image(_file):

    # Initialize COM
    pythoncom.CoInitialize()
    
    # Open Excel
    excel = win32com.client.Dispatch("Excel.Application")
    excel.Visible = False  # Run in the background
    wb = excel.Workbooks.Open(_file)

    # Extract first sheet
    _sheet = excel.Sheets('Bar')
    
    shape = _sheet.Shapes[0]
    shape.Copy()
    image = ImageGrab.grabclipboard()
    # Saves the image into the existing png file (overwriting) TODO ***** Have try except?
    outputfile = f'{os.getcwd()}/preview_chart.png'
    image.save(outputfile, 'png')
            
    # Close Excel properly
    wb.Close(False)
    excel.Quit()

    # Uninitialize COM
    pythoncom.CoInitialize()

    return outputfile




def bar_graph(df):
 
    st.radio('Bar Grouping', options=['Standard', 'Stacked', 'Clustered', 'Percent Stacked'], key='bar_group', horizontal=True)
    if st.session_state['bar_group'] == 'Standard':
        _bgroup = 'standard'
    if st.session_state['bar_group'] == 'Stacked':
        _bgroup = 'stacked'
    if st.session_state['bar_group'] == 'Clustered':
        _bgroup = 'clustered'
    if st.session_state['bar_group'] == 'Percent Stacked':
        _bgroup = 'percentstacked'

    # Create a new Workbook
    wb = openpyxl.Workbook()
    wb.create_sheet('Bar')
    ws = wb['Bar']
    wb.save('Bar.xlsx')    

    l, w = df.shape

    # append data to excel sheet
    writer = pd.ExcelWriter('Bar.xlsx', engine='openpyxl')
    df.to_excel(writer, sheet_name='Bar', startrow = 0,index = False)
    writer.close()

    wb = openpyxl.load_workbook('Bar.xlsx')
    ws = wb['Bar']

    # Create a BarChart object
    _chart = BarChart()
    _chart.type = 'col'
    _chart.style = 2
    _chart.grouping = _bgroup
    
    if _bgroup == 'stacked':
        _chart.overlap = 100
        _chart.gapWidth = 10

    # _chart.x_axis.delete = False
    # _chart.y_axis.delete = False
    _chart.legend.position = 'b'

    # Define the data (sales values) for the chart
    data = Reference(ws, min_col=2, min_row=1, max_row=l+1, max_col=w)

    # Define the categories (years)
    categories = Reference(ws, min_col=1, min_row=2, max_row=l+1, max_col=1)

    # Add data and categories to the chart
    _chart.add_data(data, titles_from_data=True)
    _chart.set_categories(categories)
    _chart.shape = 4
    _chart.dataLabels = DataLabelList()
    _chart.dataLabels.showVal = True
    _chart.DataTable = DataTable()
    _chart.height = 15
    _chart.width = 30

    _chart.title = "Custom Bar Chart"
    _chart.y_axis.title = "Values"
    _chart.x_axis.title = "Categories"
    _chart.y_axis.majorGridlines = None
    _chart.x_axis.majorGridlines = None
    _chart.y_axis.tickLblPos = "low"
    _chart.x_axis.tickLblPos = "low"    

    btn_create_graph = st.button('Create Graph')
    
    if btn_create_graph:
        
        # Add the chart to the worksheet
        ws.add_chart(_chart, "J1")

        # Save the Workbook to a file
        wb.save(f'{os.getcwd()}/Bar.xlsx')
        wb.close()

        _file = f'{os.getcwd()}/Bar.xlsx'
        img_preview = convert_to_image(_file)

        st.success('Chart Preview')
        st.image(img_preview)
             
        result_file = open('Bar.xlsx', 'rb')
        st.success(f':red[NOTE:] Downloaded file will go to the :red[Downloads Folder]')
        st.download_button(label='📥 Download Cleaned Raw', data=result_file ,file_name= f'testing_graph.xlsx')   
        
        



    return