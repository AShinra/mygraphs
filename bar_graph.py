import openpyxl
import pandas as pd
import streamlit as st
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.plotarea import DataTable
from openpyxl.drawing.text import Paragraph, ParagraphProperties, CharacterProperties, Font
from streamlit_option_menu import option_menu
# import win32com.client
import PIL
from PIL import ImageGrab, Image
import os
import sys
import pythoncom
import excel2img


def convert_to_image(_file):

    # # Initialize COM
    # pythoncom.CoInitialize()
    
    # # Open Excel
    # excel = win32com.client.Dispatch("Excel.Application")
    # excel.Visible = False  # Run in the background
    # wb = excel.Workbooks.Open(_file)

    # # Extract first sheet
    # _sheet = excel.Sheets('Bar')
    
    # shape = _sheet.Shapes[0]
    # shape.Copy()
    # image = ImageGrab.grabclipboard()
    # # Saves the image into the existing png file (overwriting) TODO ***** Have try except?
    outputfile = f'{os.getcwd()}/preview_chart.png'
    # image.save(outputfile, 'png')
            
    # # Close Excel properly
    # wb.Close(False)
    # excel.Quit()

    # # Uninitialize COM
    # pythoncom.CoInitialize()

    excel2img.export_img(_file, outputfile, "ChartSheet", None)

    return outputfile


def download_excelfile(_file):

    result_file = open(_file, 'rb')
    st.success(f':red[NOTE:] Downloaded file will go to the :red[Downloads Folder]')
    st.download_button(label='📥 Download Cleaned Raw', data=result_file ,file_name= f'excel_data_and_graph.xlsx')

    return

def style_option_menu():

    _style = {
        'nav-link':{'font-size':'12px'},
        'menu-title':{'font-size':'15px'}}

    return _style


def bar_customization():

    # default values
    _bgroup = 'standard'
    _orientation = 'col'
    l_loc = 't'
    
    c_style = st.number_input('Chart Style', min_value=1, max_value=100, step=1)

    colt1, colt2, colt3 = st.columns(3)
    
    with colt1:    
        c_title = st.text_input('Chart Title')
        if c_title == '':
            c_title = 'Bar Chart Title'
    
    with colt2:
        x_title = st.text_input('X Axis Title')
        if x_title == '':
            x_title = 'Categories'

    with colt3:
        y_title = st.text_input('Y Axis Title')
        if y_title == '':
            y_title = 'Values'
            
    col1, col2, col3 = st.columns(3)
    with col1:
        with st.expander('Bar Grouping'):
            # Customization for Bar Groupings
            grouping = option_menu(
                menu_title='',
                options=['Standard', 'Stacked', 'Clustered', 'Percent Stacked'],
                orientation='vertical',
                key='bar_group',
                default_index=0,
                styles=style_option_menu())
            
            if st.session_state['bar_group'] == 'Standard':
                _bgroup = 'standard'
            if st.session_state['bar_group'] == 'Stacked':
                _bgroup = 'stacked'
            if st.session_state['bar_group'] == 'Clustered':
                _bgroup = 'clustered'
            if st.session_state['bar_group'] == 'Percent Stacked':
                _bgroup = 'percentstacked'
    with col2:
        with st.expander('Bar Orientation'):   
            # Customization for Bar Orientation
            bar_orientation = option_menu(
                menu_title='',
                options=['Vertical', 'Horizontal'],
                orientation='vertical',
                default_index=0,
                styles=style_option_menu())

            if bar_orientation=='Vertical':
                _orientation = 'col'
            if bar_orientation=='Horizontal':
                _orientation = 'bar'
    with col3:  
        with st.expander('Legend Location'):   
            # Customization for Bar Orientation
            legend_loc = option_menu(
                menu_title='',
                options=['Top Right', 'Top', 'Bottom', 'Right', 'Left'],
                orientation='vertical',
                styles=style_option_menu())

            if legend_loc=='Top Right':
                l_loc = 'tr'
            if legend_loc=='Top':
                l_loc = 't'
            if legend_loc=='Right':
                l_loc = 'r'
            if legend_loc=='Left':
                l_loc = 'l'
            if legend_loc=='Bottom':
                l_loc = 'b'
    
    cold1, cold2, cold3 = st.columns(3)

    with cold1:
        with st.expander('Data Labels'):
            # Customization for Data Labels
            if st.checkbox('Show Value'):
                dlabels = True
            else:
                dlabels = False
    
    with cold2:
        with st.expander('Data Table'):
        # Customization for Data Labels
            if st.checkbox('Show Table'):
                dtable = True
            else:
                dtable = False
    
    with cold3:
        with st.expander('Grid Lines'):
        # Customization of Grid Lines
            if st.checkbox('Show X Major'):
                xmajor = True
            else:
                xmajor = False

            if st.checkbox('Show Y Major'):
                ymajor = True
            else:
                ymajor = False 
        
    return c_style, c_title, x_title, y_title, _bgroup, _orientation, l_loc, dlabels, dtable, xmajor, ymajor





def bar_graph(df):
    
    c_style, c_title, x_title, y_title, _bgroup, _orientation, l_loc, dlabels, dtable, xmajor, ymajor = bar_customization()

    # Create a new Workbook
    wb = openpyxl.Workbook()
    wb.create_sheet('Bar')
    ws = wb['Bar']
    wb.save('Bar.xlsx')    

    l, w = df.shape

    # append data to excel sheet
    writer = pd.ExcelWriter('Bar.xlsx', engine='openpyxl')
    df.to_excel(writer, sheet_name='Bar', startrow = 0,index = False)
    writer.close()

    wb = openpyxl.load_workbook('Bar.xlsx')
    ws = wb['Bar']

    # Create a BarChart object
    _chart = BarChart()
    _chart.type = _orientation
    _chart.style = c_style
    _chart.grouping = _bgroup
    
    if _bgroup == 'stacked':
        _chart.overlap = 100
        _chart.gapWidth = 10

    # _chart.x_axis.delete = False
    # _chart.y_axis.delete = False
    _chart.legend.position = l_loc

    # Define the data (sales values) for the chart
    data = Reference(ws, min_col=2, min_row=1, max_row=l+1, max_col=w)

    # Define the categories (years)
    categories = Reference(ws, min_col=1, min_row=2, max_row=l+1, max_col=1)

    # Add data and categories to the chart
    _chart.add_data(data, titles_from_data=True)
    _chart.set_categories(categories)

    if dlabels == True:
        _chart.dataLabels = DataLabelList()
        _chart.dataLabels.showVal = True    

    if dtable == True:
        _chart.plot_area.dTable = DataTable(True, True, True, True)

    _chart.height = 15
    _chart.width = 30

    # chart title
    _chart.title = c_title

    title_font = CharacterProperties(
        latin=Font(typeface="Calibri"),  # Font type
        sz=1800,  # Font size (18pt, in 1/100 pt)
        b=True,  # Bold
        solidFill="FF0000"  # Red color
    )

    # Apply text properties
    p = Paragraph(
        pPr=ParagraphProperties(),
        endParaRPr=title_font
    )
    _chart.title.tx.rich.p.append(p)



    _chart.y_axis.title = y_title
    _chart.x_axis.title = x_title

    if xmajor == False:
        _chart.x_axis.majorGridlines = None
    
    if ymajor == False:
        _chart.y_axis.majorGridlines = None
    

    _chart.y_axis.tickLblPos = "low"
    _chart.x_axis.tickLblPos = "low"    

    btn_create_graph = st.button('Create Graph')
    
    if btn_create_graph:
        
        # Add the chart to the worksheet
        ws.add_chart(_chart, "J1")

        # Save the Workbook to a file
        wb.save(f'{os.getcwd()}/Bar.xlsx')
        wb.close()

        _file = f'{os.getcwd()}/Bar.xlsx'
        img_preview = convert_to_image(_file)

        with st.expander('Chart Preview'):
            st.image(img_preview)

        download_excelfile(_file)
             
        
        



    return